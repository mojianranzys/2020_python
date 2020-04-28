#-*- coding: UTF-8 -*-

import sys
import time
import urllib
import urllib2
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook
import io

#sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='gb18030') 

reload(sys)
sys.setdefaultencoding('utf8')

#Some User Agents
hds={'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:74.0) Gecko/20100101 Firefox/74.0'}
print hds

def book_spider(book_tag):
    page_num=0;
    book_list=[]
    try_times=0
    # 测试脚本，设置输出限制，如果想输出全部，则while(1):
    while(page_num <= 2):
        #url='http://www.douban.com/tag/%E5%B0%8F%E8%AF%B4/book?start=0' # For Test
        #urllib.quote:使用％xx转义符替换字符串中的特殊字符
        url='http://www.douban.com/tag/'+urllib.quote(book_tag)+'/book?start='+str(page_num*15) #每一页有15本书
        print(url)
        time.sleep(np.random.rand()*5)
        
        #Last Version
        try:
            # request:GET 
            # urllib.request.urlopen  
            req = urllib2.Request(url, headers=hds)
            #print(req)
            source_code = urllib2.urlopen(req).read()
            plain_text=str(source_code)
            #print(plain_text)   
        except (urllib2.HTTPError, urllib2.URLError), e:
            print e
            continue
  
        #BeautifulSoup:将复杂的html文档转换为树形结构，每一个节点都是一个对象,使用lxml解释器，速度更快
        soup = BeautifulSoup(plain_text,"lxml")
        #print(soup)
        # <div>:定义文档中的节
        # mod book-list：每一页中出现一次 表示显示书的列表
        list_soup = soup.find('div', {'class': 'mod book-list'})
        #print(list_soup)
        try_times+=1;
        if list_soup==None and try_times<5:
            continue
        elif list_soup==None or len(list_soup)<=1:
            break # Break when no informatoin got after 200 times requesting
        
        # <dd>:定义定义列表中项目的描述
        # <a>:定义超文本链接
        # class是调用css样式的属性,desc是属性名称(例如<div class="desc"> ： 王俊凯 / 新星 / 2018-8 / 88.00元)
        for book_info in list_soup.findAll('dd'):
            #string.strip()：用于移除字符串头尾指定的字符（默认为空格或换行符）或字符序列。
            title = book_info.find('a', {'class':'title'}).string.strip()
            print(title.encode('gb18030'))
            desc = book_info.find('div', {'class':'desc'}).string.strip()
            print desc.encode('gb18030')
            desc_list = desc.split('/')
            #print(desc_list)
            book_url = book_info.find('a', {'class':'title'}).get('href')
            print(book_url)
            
            try:
                author_info = '作者/译者： ' + '/'.join(desc_list[0:-3])
            except:
                author_info ='作者/译者： 暂无'
            try:
                pub_info = '出版信息： ' + '/'.join(desc_list[-3:])
            except:
                pub_info = '出版信息： 暂无'
            try:
                rating = book_info.find('span', {'class':'rating_nums'}).string.strip()
            except:
                rating='0.0'
            try:
                people_num = get_people_num(book_url)
                #people_num = people_num.strip('人评价')
            except:
                people_num ='0'
            print(author_info.encode('gb18030'))
            print(pub_info.encode('gb18030'))
            print(rating.encode('gb18030'))
            print(people_num)
            book_list.append([title,rating,people_num,author_info,pub_info])
            #print(book_list)
            try_times=0 #set 0 when got valid information
        page_num+=1
        print 'Downloading Information From Page %d' % page_num
    return book_list

# 评价的人数：
def get_people_num(url):
    #url='http://book.douban.com/subject/6082808/?from=tag_all' # For Test
    try:
        req = urllib2.Request(url, headers=hds)
        source_code = urllib2.urlopen(req).read()
        plain_text=str(source_code)   
    except (urllib2.HTTPError, urllib2.URLError), e:
        print e
    soup = BeautifulSoup(plain_text,"lxml")
    people_num=soup.find('div',{'class':'rating_sum'}).findAll('span')[1].string.strip()
    return people_num

def do_spider(book_tag_lists):
    book_lists=[]
    for book_tag in book_tag_lists:
        book_list=book_spider(book_tag)
        book_list=sorted(book_list,key=lambda x:x[1],reverse=True)
        book_lists.append(book_list)
    return book_lists
    # sorted:可对所有可迭代的对象进行排序操作，返回的是一个新的list,reverse=T(降序)，reverse=F(升序：默认)
    # key=lambda x:x[1]：lambda是一个隐函数，是固定写法，不要写成别的单词；
    #           x表示列表中的一个元素，在这里，表示一个元组，x只是临时起的一个名字，可以使用任意的名字；x[1]表示元组里的第二个元素

def print_book_lists_excel(book_lists,book_tag_lists):
    wb=Workbook(write_only = True) # 打开一个只写入的工作簿
    ws=[]
    for i in range(len(book_tag_lists)):
        ws.append(wb.create_sheet(title=book_tag_lists[i].decode())) #创建sheet
    for i in range(len(book_tag_lists)): 
        ws[i].append(['序号','书名','评分','评价人数','作者','出版社'])   #插入表头
        count=1
        for bl in book_lists[i]:
            ws[i].append([count,bl[0],float(bl[1]),int(bl[2]),bl[3],bl[4]])
            count+=1
    save_path='book_list'
    for i in range(len(book_tag_lists)):
        save_path+=('-'+book_tag_lists[i].decode())
    save_path+='.xlsx'
    wb.save(save_path) #未设置保存路径，即保存在当前路径下D:\learn_python.




if __name__=='__main__':
    #book_tag_lists = ['心理','判断与决策','算法','数据结构','经济','历史']
    #book_tag_lists = ['传记','哲学','编程','创业','理财','社会学','佛教']
    #book_tag_lists = ['思想','科技','科学','web','股票','爱情','两性']
    #book_tag_lists = ['计算机','机器学习','linux','android','数据库','互联网']
    #book_tag_lists = ['数学']
    #book_tag_lists = ['摄影','设计','音乐','旅行','教育','成长','情感','育儿','健康','养生']
    #book_tag_lists = ['商业','理财','管理']  
    #book_tag_lists = ['名著']
    #book_tag_lists = ['科普','经典','生活','心灵','文学']
    #book_tag_lists = ['科幻','思维','金融']
    #book_tag_lists=['个人管理','时间管理','投资','文化','宗教']
    book_tag_lists = ['摄影','音乐','旅行']
    book_lists=do_spider(book_tag_lists)
    print_book_lists_excel(book_lists,book_tag_lists)